import openpyxl
import helpers.dictionary_builder as db

def create_object_from_excel(schema_type: str, excel_file: str):

    wb = openpyxl.load_workbook(excel_file)

    tables = db.Dictionary()
    for sheet in wb.worksheets:
        if sheet.title == "NOTES":
            continue

        columnList = db.ColumnList()

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=7):

            if row[3].value is not None:
                attr = db.ColAttributes(row[2].value, row[1].value, row[3].value)
            else:
                attr = db.ColAttributes(row[2].value, row[1].value)

            col = db.Column(row[0].value, attr)

            columnList.append(col)

        parents = db.ParentList()

        for row in sheet.iter_cols(min_col=3, max_col=3, min_row=2):
            parent = row[0].value
            if parent is not None:
                parents.add_parent(parent)

            table = db.Table(sheet['A2'].value, sheet['B2'].value, columnList, parents)
            tables.add_table(sheet.title, table)

    return db.DataDictionary(schema_type, tables)
