import openpyxl
import helpers.dictionary_builder as db

def create_object_from_excel(schema_type: str, excel_file: str):

    try:
        wb = openpyxl.load_workbook(excel_file)
    except PermissionError as err:
        print("\nPermission Error. Most likely you have the workbook open and stored on OneDrive.\nEither close or store in a non-OneDrive folder.\nError was:\n{0}".format(err))
        raise
    except:
        raise
    
    tables = db.Dictionary()
    
    for sheet in wb.worksheets:
        if sheet.title == "NOTES":
            continue

        for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=30):
            for i in range(30):
                if row[i].value == "col_name":
                    col_name_idx = i 
                if row[i].value == "col_type":
                    col_type_idx = i 
                if row[i].value == "col_call":
                    col_call_idx = i 
                if row[i].value == "args":
                    args_idx = i 
                if row[i].value == "filename":
                    filename_idx = i
                if row[i].value == "rec_type":
                    rec_type_idx = i
                if row[i].value == "max_recs":
                    max_recs_idx = i
                if row[i].value == "parent":
                    parent_idx = i
                if row[i].value == "header":
                    header_idx = i
                if row[i].value == "trailer":
                    trailer_idx = i
                    
        first_row = True

        parents = db.ParentList()
        columnList = db.ColumnList()
        
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=30):
            if row[filename_idx].value is not None:
                if first_row is False:
                    if cur_parent is not None:
                        parents.add_parent(cur_parent)
                    table = db.Table(cur_rec_type, cur_max_recs, columnList, parents, cur_header, cur_trailer)
                    tables.add_table(cur_filename, table)
                    parents = db.ParentList()
                    columnList = db.ColumnList()
                else:
                    first_row = False

                cur_filename=row[filename_idx].value
                cur_rec_type=row[rec_type_idx].value
                cur_max_recs=row[max_recs_idx].value
                cur_parent=row[parent_idx].value
                cur_header=row[header_idx].value
                cur_trailer=row[trailer_idx].value
               
               
            col_call = row[col_call_idx].value
            col_type = row[col_type_idx].value
            col_args = row[args_idx].value
            col_name = row[col_name_idx].value
            
            if col_call is None:
                error_msg = "No faker call (col_call) specified for [{0}] on sheet [{1}].".format(col_name, sheet.title)
                raise ValueError(error_msg)
            
            if col_args is not None:
                attr = db.ColAttributes(col_call, col_type, col_args)
            else:
                attr = db.ColAttributes(col_call, col_type)

            col = db.Column(col_name, attr)

            columnList.append(col)

        if cur_parent is not None:
            parents.add_parent(cur_parent)
        table = db.Table(cur_rec_type, cur_max_recs, columnList, parents, cur_header, cur_trailer)
        tables.add_table(cur_filename, table)

    return db.DataDictionary(schema_type, tables)
